from openpyxl.reader.excel import load_workbook

wb=load_workbook(filename=r"C:\Users\xuhao_yang\PycharmProjects\day22\1.xlsx",read_only=True)
print(wb.get_sheet_names())
print(len(wb.get_sheet_names()))

sheetnames=wb.get_sheet_names()
ws=wb.get_sheet_by_name(sheetnames[0])

print(ws.title)
print(ws.max_row)
print(ws.max_column)


for rline in range(1,ws.max_row+1):
    for column in range(1,ws.max_column+1):
        w1=ws.cell(row=rline,column=column).value
        print(w1,end=" ")
    print("\t")