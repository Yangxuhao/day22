from openpyxl.reader.excel import load_workbook
class readxlsx():
    def __init__(self,path):
        self.path=path
        self.alltext=""
        self.wb=None
    def read(self):
        self.wb = load_workbook(filename=self.path, read_only=True)
        if self.wb!=None:
            sheetnames = self.wb.get_sheet_names()
            for i in sheetnames:
                ws = self.wb.get_sheet_by_name(i)
                for rline in range(1, ws.max_row + 1):
                    for column in range(1, ws.max_column + 1):
                        w1 = ws.cell(row=rline, column=column).value
                        print(w1, end=" ")
                    print("\t")
path=r"C:\Users\xuhao_yang\PycharmProjects\day22\1.xlsx"
read=readxlsx(path).read()